import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl,xlrd
import pathlib
from openpyxl import workbook

#setando aparencia padrao do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("green")

class App(ctk.CTk):
    def __init__(self):
         super().__init__()
         self.layout_config()
         self.appearence()
         self.todo_sistema()
    0
    def layout_config(self):
        self.title("Sistema de Gestão de Clientes.")
        self.geometry("700x500")
     
    

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self,text="Tema",bg_color="transparent",text_color=['#000','#fff']).place(x=50,y=430)
        self.opt_apm = ctk.CTkOptionMenu(self,values=["System","Dark","Light"], command= self.change_apm).place(x=50,y=460)
    
    def todo_sistema(self):
        frame = ctk.CTkFrame(self,width=700,height=50,corner_radius=0,bg_color="teal",fg_color="teal").place(x=0,y=10)
        title = ctk.CTkLabel(frame,text = "Sistema de Gestão de Clientes",font=("Century Gothic bold",24)).place(x=160,y=10)
        span = ctk.CTkLabel(frame,text = "Por favor, preencha todos os campos do formulário!",font=("Century Gothic bold",16),text_color=["#000","#fff"]).place(x=50,y=70)
        
        #Variaveis de Texto :
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        
        
        
        
        #Entrys
        name_entry = ctk.CTkEntry(self,width=250,font=("Century Gohtic", 16), textvariable= name_value)
        contact_entry = ctk.CTkEntry(self,width=120,font=("Century Gohtic", 16),textvariable= contact_value)
        age_entry = ctk.CTkEntry(self,width=50,font=("Century Gohtic", 16),textvariable= age_value)
        address_entry = ctk.CTkEntry(self,width=250,font=("Century Gohtic", 16),textvariable= address_value)
       
       #Combobox
        gender_combobox = ctk.CTkComboBox(self,values=["Masculino","Feminino"],font=("Century Gothic bold",14))
        gender_combobox.set("Masculino")
        
        #Entrada de Observações
        obs_entry = ctk.CTkTextbox(self,width=300,height=100,font=("arial", 18), border_color="#aaa",border_width=2,fg_color="transparent")
        
        
        #labels
        lb_name = ctk.CTkLabel(self,text = "Nome Completo :",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_contact = ctk.CTkLabel(self,text = "Contato :",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_age = ctk.CTkLabel(self,text = "Idade:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_gender = ctk.CTkLabel(self,text = "Gênero:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_addres = ctk.CTkLabel(self,text = "Endereço:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        lb_observations = ctk.CTkLabel(self,text = "Observações:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        def submit():
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0,END)
            
            ficheiro = openpyxl.load_workbook("Clientes.xlsx")
            folha = ficheiro.active
            
            folha.cell(column=1,row=folha.max_row+1,value=name)
            folha.cell(column=2,row=folha.max_row,value=contact)
            folha.cell(column=3,row=folha.max_row,value=age)
            folha.cell(column=4,row=folha.max_row,value=gender)
            folha.cell(column=5,row=folha.max_row,value=address)
            folha.cell(column=6,row=folha.max_row,value=obs)
            
            ficheiro.save(r"Clientes.xlsx")
            messagebox.showinfo("Sistema :", "Dados Salvos Com Sucesso")
            
    
        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0,END)
            
        
        
        
  
        
        
        #botões
        btn_submit = ctk.CTkButton(self,text="Salvar dados".upper(), command=submit,fg_color="#151",hover_color="#131").place(x=300,y=420)
        
        btn_clear = ctk.CTkButton(self,text="limpar dados".upper(), command=clear,fg_color="#555",hover_color="#333").place(x=450,y=420)
        
        #posicionando elementos na janela
        
        lb_name.place(x=50,y=120)
        name_entry.place(x=50,y=150)
        
        lb_contact.place(x=350,y=120)
        contact_entry.place(x=350,y=150)
        
        lb_age.place(x=500,y=120)
        age_entry.place(x=500,y=150)
        
        lb_gender.place(x= 350,y=190)
        gender_combobox.place(x=350,y=220)
        
        lb_addres.place(x=50,y=190)
        address_entry.place(x=50,y=220)
        
        
        lb_observations.place(x=50,y=280)
        obs_entry.place(x=200,y=280)
        
        
        
        
    
        
    
    
    def change_apm(self, nova_aparencia):
      ctk.set_appearance_mode(nova_aparencia)
    
    
if __name__ == "__main__":
    app = App()
    app.mainloop()